#!/usr/bin/env python3
"""
JSON抽出結果 → Excelファイル変換スクリプト

最終結果の JSON を読み込み、Excel (.xlsx) に出力する。
- 案件ごとに行（1行 = 1案件）
- フィールドごとに列
- シートはファイルごとに分割 + 全体サマリーシート
"""

import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl がインストールされていません。pip install openpyxl を実行してください。")
    sys.exit(1)


# ── スタイル定義 ──
HEADER_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="Yu Gothic", size=10)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def flatten_case(case: dict) -> dict:
    """ネストした案件データをフラットな dict に変換。"""
    loc = case.get("location") or {}
    if isinstance(loc, str):
        loc = {"city": loc}
    rate = case.get("rate") or {}
    if isinstance(rate, (int, float, str)):
        rate = {"min": rate}
    period = case.get("period") or {}
    if isinstance(period, str):
        period = {"note": period}
    trade = case.get("trade_flow") or {}
    if isinstance(trade, str):
        trade = {"contract_type_jp": trade}
    jl = case.get("japanese_level") or {}
    if isinstance(jl, str):
        jl = {"level": jl}
    el = case.get("english_level") or {}
    if isinstance(el, str):
        el = {"level": el}
    wh = case.get("working_hours") or {}
    if isinstance(wh, str):
        wh = {"overtime": wh}
    src = case.get("source") or {}
    if isinstance(src, str):
        src = {"filename": src}
    exp = case.get("experience_years") or {}
    if isinstance(exp, (int, float, str)):
        exp = {"description": str(exp)}

    def skill_str(skills):
        if not skills:
            return ""
        return "\n".join(s.strip() for s in skills if s.strip())

    def remote_str(policy):
        mapping = {
            "full_remote": "フルリモート",
            "hybrid": "ハイブリッド（在宅+出社）",
            "office_only": "オンサイトのみ",
            "not_specified": "",
        }
        return mapping.get(policy, policy or "")

    def jl_str(level):
        mapping = {
            "native": "ネイティブ",
            "business": "ビジネスレベル",
            "n2": "N2以上",
            "n3": "N3以上",
            "n4": "N4以上",
            "not_specified": "",
        }
        return mapping.get(level, level or "")

    # Fix dates that are before current year (LLM hallucinations)
    now = datetime.now()
    current_year = now.year
    current_month = now.month

    def fix_date(date_str):
        """Fix date: if year < current_year, bump to current_year (recruitment is always future-looking)."""
        if not date_str or not isinstance(date_str, str):
            return date_str
        try:
            parts = date_str.split("-")
            y = int(parts[0])
            if y < current_year:
                m = int(parts[1]) if len(parts) > 1 else 1
                d = int(parts[2]) if len(parts) > 2 else 1
                target_year = current_year if m >= current_month else current_year + 1
                return f"{target_year}-{m:02d}-{d:02d}"
        except (ValueError, IndexError):
            pass
        return date_str

    # 即日参画で開始日未設定 → 本日付
    start_date = period.get("start_date") or ""
    if not start_date and case.get("immediate_start"):
        start_date = now.strftime("%Y-%m-%d")
    start_date = fix_date(start_date)
    end_date = fix_date(period.get("end_date") or "")

    return {
        "案件名": case.get("project_name") or "",
        "案件概要": case.get("project_description") or "",
        "必須スキル": skill_str(case.get("skill_requirement")),
        "歓迎スキル": skill_str(case.get("preferred_skills")),
        "経験年数(最小)": exp.get("min") or "",
        "経験年数(最大)": exp.get("max") or "",
        "経験年数(詳細)": exp.get("description") or "",
        "勤務地(都市)": loc.get("city") or "",
        "最寄駅": loc.get("station") or "",
        "リモート方針": remote_str(loc.get("remote_policy")),
        "リモート詳細": loc.get("remote_detail") or "",
        "単価(下限)": rate.get("min") or "",
        "単価(上限)": rate.get("max") or "",
        "単価単位": rate.get("unit") or "",
        "通貨": rate.get("currency") or "JPY",
        "単価備考": rate.get("note") or "",
        "開始日": start_date,
        "終了日": end_date,
        "予定月数": period.get("duration_months") or "",
        "長期フラグ": "✓" if period.get("long_term") else "",
        "期間備考": period.get("note") or "",
        "募集人数": case.get("headcount") or "",
        "業種": case.get("industry") or "",
        "契約形態": trade.get("contract_type_jp") or trade.get("contract_type") or "",
        "商流階層": trade.get("layers") or "",
        "最終顧客": trade.get("end_client") or "",
        "日本語レベル": jl_str(jl.get("level")),
        "日本語詳細": jl.get("detail") or "",
        "英語レベル": el.get("level") or "",
        "英語詳細": el.get("detail") or "",
        "勤務開始": wh.get("start") or "",
        "勤務終了": wh.get("end") or "",
        "フレックス": "✓" if wh.get("flex_time") else "",
        "残業": wh.get("overtime") or "",
        "面接回数": case.get("interviews") or "",
        "即日参画": "✓" if case.get("immediate_start") else "",
        "選考フロー": case.get("screening_flow") or "",
        "備考": case.get("remarks") or "",
        "ソース形式": src.get("original_format") or "",
        "ファイル名": src.get("filename") or "",
        "案件全文": (case.get("original_text") or "")[:30000],
    }


def write_sheet(ws, cases: list[dict], sheet_title: str):
    """案件リストをシートに書き込む。"""
    if not cases:
        ws.cell(row=1, column=1, value="案件データがありません")
        return

    # Convert all cases to flat dicts
    flat_cases = [flatten_case(c) for c in cases]

    # Columns
    headers = list(flat_cases[0].keys())
    col_count = len(headers)

    # ── ヘッダー行 ──
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # ── データ行 ──
    for row_idx, flat in enumerate(flat_cases, 2):
        for col_idx, header in enumerate(headers, 1):
            value = flat.get(header, "")
            # 数値変換（空文字以外の数値文字列）
            if isinstance(value, str) and value.strip():
                try:
                    value = int(value)
                except ValueError:
                    try:
                        value = float(value)
                    except ValueError:
                        pass
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER
            # 偶数行は背景色
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # ── 列幅の自動調整（概算） ──
    col_widths = {
        "案件名": 40,
        "案件概要": 50,
        "必須スキル": 50,
        "歓迎スキル": 40,
        "勤務地(都市)": 18,
        "リモート方針": 20,
        "単価(下限)": 10,
        "単価(上限)": 10,
        "単価単位": 10,
        "開始日": 12,
        "終了日": 12,
        "募集人数": 10,
        "業種": 12,
        "契約形態": 15,
        "日本語レベル": 14,
        "面接回数": 10,
        "備考": 40,
        "ファイル名": 30,
    }
    for col_idx, header in enumerate(headers, 1):
        width = col_widths.get(header, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 行の高さ（スキルが複数行の場合に対応）
    for row_idx in range(2, len(flat_cases) + 2):
        skill_cell = ws.cell(row=row_idx, column=headers.index("必須スキル") + 1)
        if skill_cell.value and isinstance(skill_cell.value, str):
            line_count = skill_cell.value.count("\n") + 1
            if line_count > 1:
                ws.row_dimensions[row_idx].height = max(15, line_count * 16)

    # フリーズペイン（ヘッダー固定）
    ws.freeze_panes = "A2"

    # オートフィルター
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}{len(flat_cases) + 1}"


def main():
    data_dir = Path(__file__).parent.parent / "data" / "output"

    # 最新の最終結果ファイルを検索
    import glob
    final_files = sorted(glob.glob(str(data_dir / "final_result_*.json")))
    if not final_files:
        # フォールバック: LLM結果 or ルール結果
        llm_files = sorted(glob.glob(str(data_dir / "llm_extraction_result_*.json")))
        rule_files = sorted(glob.glob(str(data_dir / "extraction_result_*.json")))
        target = llm_files[-1] if llm_files else (rule_files[-1] if rule_files else None)
        if not target:
            print("変換対象のJSONファイルが見つかりません")
            sys.exit(1)
        print(f"最終結果なし → LLM結果を使用: {target}")
    else:
        target = final_files[-1]
        print(f"最終結果を使用: {target}")

    with open(target, encoding="utf-8") as f:
        data = json.load(f)

    # Excel ブック作成
    wb = openpyxl.Workbook()

    # ── 1. サマリーシート ──
    ws_summary = wb.active
    ws_summary.title = "サマリー"
    stats = data.get("stats", {})

    summary_data = [
        ["項目", "値"],
        ["抽出日時", data.get("extraction_date", "")],
        ["全案件数", stats.get("total_cases", 0)],
        ["処理ファイル数", stats.get("files_processed", 0)],
        ["抽出モード", stats.get("extraction_mode", "")],
        ["LLMモデル", stats.get("model", "")],
    ]

    for row_idx, (key, val) in enumerate(summary_data, 1):
        cell_k = ws_summary.cell(row=row_idx, column=1, value=key)
        cell_v = ws_summary.cell(row=row_idx, column=2, value=val)
        if row_idx == 1:
            cell_k.font = HEADER_FONT
            cell_k.fill = HEADER_FILL
            cell_v.font = HEADER_FONT
            cell_v.fill = HEADER_FILL
        else:
            cell_k.font = Font(name="Yu Gothic", bold=True, size=10)
            cell_v.font = CELL_FONT
        cell_k.border = THIN_BORDER
        cell_v.border = THIN_BORDER

    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 40

    # ファイル別内訳
    results = data.get("results", {})
    row_start = len(summary_data) + 2
    ws_summary.cell(row=row_start, column=1, value="ファイル別内訳").font = Font(name="Yu Gothic", bold=True, size=11)
    ws_summary.cell(row=row_start + 1, column=1, value="ファイル名").font = HEADER_FONT
    ws_summary.cell(row=row_start + 1, column=1).fill = HEADER_FILL
    ws_summary.cell(row=row_start + 1, column=2, value="案件数").font = HEADER_FONT
    ws_summary.cell(row=row_start + 1, column=2).fill = HEADER_FILL

    for i, (fname, result) in enumerate(results.items(), 2):
        ws_summary.cell(row=row_start + i, column=1, value=fname).font = CELL_FONT
        ws_summary.cell(row=row_start + i, column=1).border = THIN_BORDER
        ws_summary.cell(row=row_start + i, column=2, value=result.get("case_count", 0)).font = CELL_FONT
        ws_summary.cell(row=row_start + i, column=2).border = THIN_BORDER

    # ── 2-4. ファイル別シート ──
    for fname, result in results.items():
        cases = result.get("cases", [])
        if not cases:
            continue

        # シート名 (ファイル名の先頭31文字以内 / 禁則文字除去)
        sheet_name = fname
        for ch in ['\\', '/', '*', '?', ':', '[', ']']:
            sheet_name = sheet_name.replace(ch, '_')
        sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, cases, sheet_name)
        print(f"  シート「{sheet_name}」: {len(cases)} 件書き込み")

    # ── 5. 全案件一括シート ──
    all_cases = []
    for result in results.values():
        all_cases.extend(result.get("cases", []))

    if all_cases:
        ws_all = wb.create_sheet(title="全案件一覧", index=0)  # 先頭に挿入
        write_sheet(ws_all, all_cases, "全案件一覧")
        print(f"  シート「全案件一覧」: {len(all_cases)} 件書き込み")

    # ── 保存 ──
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = data_dir / f"jp_recruit_cases_{timestamp}.xlsx"
    wb.save(str(output_path))

    print(f"\n✅ Excel出力完了: {output_path}")
    print(f"   シート構成:")
    for ws in wb.worksheets:
        print(f"   - {ws.title}: {ws.max_row - 1} 行 × {ws.max_column} 列")


if __name__ == "__main__":
    main()
